import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import filedialog
import os


def sheets_to_one():
    """
    parameters: none
    returns: none
    description: takes a single xlsx file and appends the sheets vertically into a single workbook/sheet
    """
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)  # This line will make the dialog appear on top
    file_path = filedialog.askopenfilename()

    old = openpyxl.load_workbook(file_path, data_only=True)
    new = openpyxl.Workbook()
    cell_range = input("What cell range do you want to pull from? (ex. A1:O100)\n")

    all_data = []
    transposeq = input("Do you want to transpose the data? (y/n)\n")
    ignoreq = input("Do you want to drop empty rows? (y/n)\n")
    for sheet in old.worksheets:
        data = sheet[cell_range]
        rows_list = []
        for row in data:
            if ignoreq == 'y':
                if not all(cell.value is None for cell in row):
                    rows_list.append([cell.value for cell in row])
            else:
                rows_list.append([cell.value for cell in row])
        if transposeq == 'y':  
            df = pd.DataFrame(rows_list).transpose()
        else:
            df = pd.DataFrame(rows_list)
        all_data.append(df)
    final_df = pd.concat(all_data)
    final_df.to_excel(file_path + "_Spreadsheet_Tools.xlsx", index=False)


def books_to_one():
    """
    parameters: none
    returns: none
    description: takes a folder of xlsx files and appends the sheets vertically into a single workbook/sheet
    """
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    folder_path = filedialog.askdirectory()
    files = os.listdir(folder_path)
    files_xlsx = [f for f in files if f[-4:] == 'xlsx']
    wb = openpyxl.load_workbook(folder_path + "/" + files_xlsx[0], data_only=True)
    ws = wb.active
    ws.title = "All Data"
    ignoreq = input("Do you want to drop empty rows? (y/n)\n")
    for file in files_xlsx[1:]:
        workbook = openpyxl.load_workbook(folder_path + "/" + file, data_only=True)
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                if ignoreq == 'y':
                    if not all(cell.value is None for cell in row):
                        values = [cell.value for cell in row]
                        ws.append(values)
                else:
                    values = [cell.value for cell in row]
                    ws.append(values)
    wb.save(folder_path + "/" + "All_Data.xlsx")
    wb.close()
    workbook.close()


def help(option):
    if option == '2':
        print("Many books to one takes a folder of xlsx files and appends the sheets vertically into a single workbook/sheet\nWarning only pulls the first sheet.")
    if option == '1':
        print("Many sheets to one takes a specific range and takes all sheets in the workbook at vertically appends them all into a single new workbook/sheet. There is an option to transpose the data")
    else:
        print("Invalid option")

def main():
    print("""

   _____                          _     _               _     _______          _ 
  / ____|                        | |   | |             | |   |__   __|        | |
 | (___  _ __  _ __ ___  __ _  __| |___| |__   ___  ___| |_     | | ___   ___ | |
  \___ \| '_ \| '__/ _ \/ _` |/ _` / __| '_ \ / _ \/ _ \ __|    | |/ _ \ / _ \| |
  ____) | |_) | | |  __/ (_| | (_| \__ \ | | |  __/  __/ |_     | | (_) | (_) | |
 |_____/| .__/|_|  \___|\__,_|\__,_|___/_| |_|\___|\___|\__|    |_|\___/ \___/|_|
        | |                                                                      
        |_|                                                                      
   
        """)
    choice = input("What would you like to do?\n1. many sheets to one\n2. many workbooks to one\n0. exit\n\nhelp example: help 1\n\n")
    if choice.startswith('help'):
        _, option = choice.split()
        help(option)
    if choice == '1':
        sheets_to_one()
    if choice =='2':
        books_to_one()
    if choice == '0':
        os.system('cls' if os.name == 'nt' else 'clear')
        exit()
    else:
        os.system('cls' if os.name == 'nt' else 'clear')
        main()
if __name__ == "__main__":
    main()